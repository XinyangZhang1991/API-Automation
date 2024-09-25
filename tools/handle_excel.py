# Turn the testcase in Excel spreadsheet into the dictionary format  we need

from openpyxl import load_workbook
from tools.handle_path import excel_path_day16
from loguru import logger


def read_excel(filename,sheetname):
    try:
        wb = load_workbook(filename)
        if sheetname not in wb.sheetnames:
            raise KeyError(f"sheet'{sheetname}does not exist ")
        sh = wb[sheetname] #固定的格式，就得这么写
    #     convert the sheet values into a list
        case_all = list(sh.values)


        title = case_all[0] # Extract title header
        entire_testcase_list = []  #Initalise an empty list for zip dictionary later
        # iterate over the rows, starting from the second row (actual data)
        for case in case_all[1:]:  #从列表中的第二个开始,因为第一个数据是标题
            data = dict(zip(title, case)) # Create a dictionary for each row
            print(data)
            entire_testcase_list.append(data)

        return entire_testcase_list

    except FileNotFoundError:
        logger.error (f"File'{filename}'not found.")
        return []

if __name__ =='__main__':
    from tools.handle_path import excel_path_xinyangapitesting
    case_all=read_excel(excel_path_xinyangapitesting,'login')